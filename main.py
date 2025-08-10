"""
Excel -> PDF (one page per row), simple "Key: Value" list.
- Page size / orientation is copied from the FIRST PAGE of template.pdf
- NO command-line font argument: script auto-picks an Arabic-capable font
- Draws keys with a Latin font and values with mixed-script support:
    * Arabic segments are shaped (arabic-reshaper + python-bidi) and drawn with an Arabic font
    * Non‑Arabic segments are drawn with a Latin font
    * Segments are concatenated on the same line with correct widths
- Very simple wrapping to page width (left-aligned lines for simplicity)

Run:
    python main.py --excel "sample rows.xlsx"
"""

import re
import os
import argparse
from pathlib import Path
from typing import Optional, List, Tuple
from datetime import datetime, date, time

import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.colors import black
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import landscape, A4
from openpyxl.styles.numbers import is_date_format

# ---------------------- CONFIGURATION AREA ----------------------

# Directories
IMAGES_PATH = "images"
os.makedirs("output", exist_ok=True)

# Hard-coded Page size using A4-landscape width
PAGE_WIDTH, PAGE_HEIGHT = landscape(A4)
# PAGE_HEIGHT = PAGE_WIDTH * 9.0 / 16.0         # If a 16:9 page area is required

# Alternate row background colors
LIGHT_PINK = (248/255, 236/255, 252/255) 
DARK_PINK  = (232/255, 212/255, 244/255)

CELL_BORDERS = (222/255, 185/255, 252/255)

TITLE_N_CATEGORY_COLOR = (98/255, 28/255, 154/255)

# cell small horizontal padding
CELL_PAD = 6

# Cell style height/gap
cell_height = 30
row_gap = 0

# Gap between data columns
col_gap = 10

# Layout constants
margin_l, margin_r = 20, 20
margin_t, margin_b = 20, 20

        
# --- start LEFT-SIDE IMAGE ---
img_gap = 5     # gap between image block and table block
img_pad = 5      # inner padding inside the image block

# Font Sizes
KV_FONT_SIZE = 11       # Key-Value Pair Font Size
TITLE_FONT_SIZE = 30

# ONLY list columns that are required to display
FIELDS_TO_KEEP = [
    "region", "city", "hay", "pca",
    "Population",
    "Total Premisis (Census)",
    "Remaining Premisis",
    "Female %",
    "HH Avg Size",
    "Population Density",
    "ss ARPU",
    "ss Usage",
    "Income Level (Based on ARPU)",
    "Last deployment year",
    "Major City?",
    "Visual Remark",
    "NPV",
    "IRR",
    "Forecasted Uptake",
    "Actual Uptake %",
    "Coverage",
    "Wrong Classification?",
    "# of Data centers",
    "Mobile Towers",
    "Fiberized Towers",
    "Remaining Towers",
    "ISP Cost",
    "OSP Cost",
    "STC Category",
    "Mobily Category",
    "Dawiyat Category",
    "TLS Category",
    "image"
]

EXCLUSION_FIELDS = ["region", "city", "hay", "pca", "image"]

# ---------------------- CONFIGURATION AREA ----------------------

# Arabic helpers
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
except Exception:
    arabic_reshaper = None
    get_display = None

# ---------------- Fonts ----------------
LATIN_FONT = "Helvetica"
AR_FONT_NAME = "CustomArabicFont"

LATIN_FONT_BOLD = "Helvetica-Bold"
AR_FONT_BOLD_NAME = "CustomArabicFont-Bold"

BOLD_CANDIDATES = [
    "NotoNaskhArabic-Bold.ttf",
    "NotoKufiArabic-Bold.ttf",
    "Amiri-Bold.ttf",
    "GeezaPro.ttc",
    "AlBayan.ttc",
    "Arabtype.ttf",
]


FONT_SEARCH_PATHS = [
    # current folder first (drop a copy of the ttf/ttc here for best reliability)
    ".",
    # macOS
    "/System/Library/Fonts",
    "/System/Library/Fonts/Supplemental",
    "/Library/Fonts",
    os.path.expanduser("~/Library/Fonts"),
    # Linux
    "/usr/share/fonts/truetype",
    "/usr/share/fonts",
]

# Prefer fonts that include BOTH Arabic and Latin glyphs.
AR_CANDIDATES = [
    "NotoNaskhArabic-Regular.ttf",
    "NotoKufiArabic-Regular.ttf",
    "Amiri-Regular.ttf",
    "GeezaPro.ttc",     # macOS
    "AlBayan.ttc",      # macOS
]


# ---------------- Helper for reading excel sheets ----------------
_percent_re = re.compile(r'%(?!")')  # a % not inside quotes
_dec_re     = re.compile(r'0|#')     # count decimal placeholders
_group_re   = re.compile(r'[#0],[#0]{3}')  # crude grouping detector like #,##0

def _extract_currency(fmt: str) -> str:
    if "[$" in fmt:
        try:
            part = fmt.split("[$", 1)[1]
            token = part.split("]", 1)[0]
            sym = token.split("-", 1)[0]
            return sym
        except Exception:
            pass
    for s in ["$", "€", "£", "₹", "¥", "₩", "₽", "₺", "R$", "₴", "₦"]:
        if s in fmt:
            return s
    return ""


def _format_number_by_mask(value: float, fmt: str) -> str:
    # Handle percent
    if _percent_re.search(fmt):
        # count decimals before %
        dec_part = fmt.split("%")[0]
        if "." in dec_part:
            decimals = len(_dec_re.findall(dec_part.split(".", 1)[1]))
        else:
            decimals = 0
        return f"{value*100:.{decimals}f}%"

    # Handle scientific
    if "E+" in fmt.upper() or "E-" in fmt.upper():
        # count decimals after dot if present
        decimals = 0
        if "." in fmt:
            decimals = len(_dec_re.findall(fmt.split(".", 1)[1]))
        return f"{value:.{decimals}E}"

    # Handle currency & general fixed/variable decimals
    currency = _extract_currency(fmt)
    use_grouping = bool(_group_re.search(fmt))
    decimals = 0
    if "." in fmt:
        decimals = len(_dec_re.findall(fmt.split(".", 1)[1]))

    # Build base number string
    if use_grouping:
        num = f"{value:,.{decimals}f}"
    else:
        num = f"{value:.{decimals}f}"

    # For masks using '#' after decimal, we can trim trailing zeros
    if "." in fmt and "#" in fmt.split(".", 1)[1]:
        if "." in num:
            num = num.rstrip("0").rstrip(".")

    return f"{currency}{num}" if currency else num


def _cell_rich_text_to_str(cell) -> str:
    # openpyxl 3.1+: when iter_rows(..., rich_text=True), cell.rich_text is a list of runs
    runs = getattr(cell, "rich_text", None)
    if runs:
        return "".join(getattr(r, "text", "") for r in runs)
    v = cell.value
    return "" if v is None else str(v)


def read_excel_preserve_display(path, sheet=0):
    """
    Reads an Excel sheet and returns a DataFrame of strings matching Excel's displayed text:
      - Percent: uses the cell's number_format (32% / 32.5% / 32.50%)
      - Currency / grouped numbers: respects grouping and decimals
      - Scientific: respects E-format decimals
      - Dates/Times: ISO strings (YYYY-MM-DD / HH:MM:SS)
      - Text (including rich text): concatenates runs in order
      - Blanks: ''
    """
    wb = load_workbook(path, data_only=True, read_only=True, rich_text=True)
    ws = wb[wb.sheetnames[sheet] if isinstance(sheet, int) else sheet]

    rows = ws.iter_rows(values_only=False)

    # Header row (preserve as plain strings / rich text concatenation)
    hdr_cells = next(rows)
    headers = [(_cell_rich_text_to_str(c) or "").strip() for c in hdr_cells]

    out = []
    for r in rows:
        row_out = []
        for c in r:
            v = c.value
            if v is None:
                row_out.append("")
                continue

            fmt = (c.number_format or "").strip()

            # Dates / times (Excel serials already evaluated by data_only=True)
            if is_date_format(fmt) or isinstance(v, (datetime, date, time)):
                if isinstance(v, datetime):
                    row_out.append(v.strftime("%Y-%m-%d"))
                elif isinstance(v, date):
                    row_out.append(v.strftime("%Y-%m-%d"))
                elif isinstance(v, time):
                    row_out.append(v.strftime("%H:%M:%S"))
                else:
                    row_out.append(str(v))
                continue

            # Numbers (float / int): format per mask; otherwise plain numeric string
            if isinstance(v, (int, float)):
                if fmt and fmt != "General":
                    v = round(float(v), 2)
                    row_out.append(_format_number_by_mask(float(v), fmt))
                else:
                    # Plain numeric, no extra chars
                    if isinstance(v, int):
                        row_out.append(str(v))
                    else:
                        row_out.append(format(v, ".15g"))  # compact float as string
                continue

            # Text (including rich text)
            row_out.append(_cell_rich_text_to_str(c))
        out.append(row_out)

    return pd.DataFrame(out, columns=headers)


# ---------------- Arabic detection / shaping ----------------
def is_arabic_char(ch: str) -> bool:
    return (
        ("\u0600" <= ch <= "\u06FF") or
        ("\u0750" <= ch <= "\u077F") or
        ("\u08A0" <= ch <= "\u08FF") or
        ("\uFB50" <= ch <= "\uFDFF") or
        ("\uFE70" <= ch <= "\uFEFF")
    )


def is_arabic_text(s: str) -> bool:
    if not isinstance(s, str):
        return False
    return any(is_arabic_char(ch) for ch in s)


def shape_arabic(s: str) -> str:
    if not isinstance(s, str) or not s:
        return s
    if arabic_reshaper and get_display and is_arabic_text(s):
        return get_display(arabic_reshaper.reshape(s))
    return s


# ---- Bold Arabic font detection & registration ----
def _find_font_recursive(preferred_names, bases, exts=(".ttf", ".otf", ".ttc")) -> Optional[Path]:
    for base in bases:
        bp = Path(base)
        if not bp.exists():
            continue
        for root, _, files in os.walk(bp):
            for f in files:
                lf = f.lower()
                if not lf.endswith(exts):
                    continue
                for name in preferred_names:
                    if name.lower() in lf:
                        return Path(root) / f
    return None


def find_arabic_font() -> Optional[Path]:
    for base in FONT_SEARCH_PATHS:
        try:
            for cand in AR_CANDIDATES:
                p = Path(base) / cand if base != "." else Path(cand)
                if p.exists():
                    return p
        except Exception:
            continue
    return None


def register_arabic_font() -> str:
    """
    Try to find and register an Arabic-capable font.
    Returns the font name to use for Arabic runs. Falls back to LATIN_FONT if none found.
    """
    font_path = find_arabic_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont(AR_FONT_NAME, str(font_path)))
            return AR_FONT_NAME
        except Exception:
            pass
    # Fallback (won't shape Arabic correctly if glyphs missing)
    return LATIN_FONT


# ---------------- Layout helpers ----------------
def text_width(text: str, font_name: str, font_size: float) -> float:
    return pdfmetrics.stringWidth(text, font_name, font_size)


def segment_by_script(s: str) -> List[Tuple[bool, str]]:
    """
    Split string into runs of (is_arabic, substring).
    """
    if not s:
        return []
    runs: List[Tuple[bool, str]] = []
    cur_is_ar = is_arabic_char(s[0])
    cur = [s[0]]
    for ch in s[1:]:
        ia = is_arabic_char(ch)
        if ia == cur_is_ar:
            cur.append(ch)
        else:
            runs.append((cur_is_ar, "".join(cur)))
            cur = [ch]
            cur_is_ar = ia
    runs.append((cur_is_ar, "".join(cur)))
    return runs


def draw_mixed_line(
    cnv: rl_canvas.Canvas,
    x_left: float,
    y_baseline: float,
    max_width: float,
    latin_font: str,
    ar_font: str,
    kv_text: str,
    kv_size: float,
    leading: float
) -> float:
    """
    Draw a possibly mixed-script line with simple wrapping:
      - Segment into Arabic / non-Arabic runs.
      - Shape Arabic runs.
      - Place runs left-to-right, switching fonts per run.
    Returns the next y-baseline.
    """
    # Wrap manually: build lines that fit width
    lines: List[List[Tuple[str, str]]] = []  # list of lines; each line is list of (font_name, segment_text)
    current_line: List[Tuple[str, str]] = []
    current_width = 0.0

    # Split by spaces to allow wrapping at spaces, but keep script segmentation inside tokens
    words = kv_text.split(" ")
    for wi, w in enumerate(words):
        w_runs = segment_by_script(w)
        # shape / set font per run and compute total word width (including a space before it if needed)
        word_segments: List[Tuple[str, str]] = []
        word_width = 0.0

        # space prefix if not first word of line
        space_text = (" " if (current_line or wi > 0) else "")
        if space_text:
            word_segments.append((latin_font, space_text))
            word_width += text_width(space_text, latin_font, kv_size)

        for is_ar, seg in w_runs:
            seg2 = shape_arabic(seg) if is_ar else seg
            font = ar_font if is_ar else latin_font
            word_segments.append((font, seg2))
            word_width += text_width(seg2, font, kv_size)

        # If this word doesn't fit, break line (unless it's the first on the line)
        if current_line and current_width + word_width > max_width:
            lines.append(current_line)
            current_line = []
            current_width = 0.0
            # Add the word (without a leading space on a new line)
            # remove the space segment if present at start
            if word_segments and word_segments[0][1] == " ":
                word_segments = word_segments[1:]
                word_width = sum(text_width(seg, fnt, kv_size) for fnt, seg in word_segments)
        current_line.extend(word_segments)
        current_width += word_width

    if current_line:
        lines.append(current_line)

    # Draw lines
    y = y_baseline
    for line in lines:
        x = x_left
        for font, seg in line:
            cnv.setFont(font, kv_size)
            cnv.drawString(x, y, seg)
            x += text_width(seg, font, kv_size)
        y -= leading
    return y


def _normalize_value(v):
    """Format values to fit nicely inside cells."""
    if pd.isna(v) or v is None or str(v).strip() == "":
        return "—"
    if isinstance(v, (int,)) and not isinstance(v, bool):
        return f"{v}"
    # try number
    try:
        f = float(v)
        # trim trailing zeros, max 3 decimals
        s = f"{f:.3f}".rstrip("0").rstrip(".")
        return s
    except Exception:
        pass
    # booleans / yes-no
    sv = str(v).strip()
    if sv.lower() in {"true", "yes", "y", "1"}:
        return "Yes"
    if sv.lower() in {"false", "no", "n", "0"}:
        return "No"
    return sv


def draw_kv_row(cnv, x, y, row_height, key_text, val_text, key_w, val_w, font_latin, font_ar, font_size, row_index):
    """Draw one table row with adjacent KEY and VALUE cells."""

    # Use row index to pick color — assumes caller passes `row_index` param
    bg_color = LIGHT_PINK if row_index % 2 == 0 else DARK_PINK
    cnv.setFillColorRGB(*bg_color)

    # key cell background
    cnv.rect(x, y - row_height, key_w, row_height, stroke=0, fill=1)
    # value cell background
    cnv.rect(x + key_w, y - row_height, val_w, row_height, stroke=0, fill=1)

    # cell borders over the fill
    cnv.setStrokeColorRGB(*CELL_BORDERS)
    # rgb(222,185,252)
    cnv.setLineWidth(0.5)
    cnv.rect(x, y - row_height, key_w, row_height, stroke=1, fill=0)
    cnv.rect(x + key_w, y - row_height, val_w, row_height, stroke=1, fill=0)

    # vertically center single-line text
    baseline = y - (row_height - font_size) / 2 - 2.5
    
    # Set key text bold and color
    cnv.setFillColorRGB(*TITLE_N_CATEGORY_COLOR)
    cnv.setFont(font_latin + "-Bold", font_size)

    # key (left cell) — bold + purple
    cnv.setFillColorRGB(*TITLE_N_CATEGORY_COLOR)  # #621c9a
    draw_mixed_line(
        cnv=cnv,
        x_left=x + CELL_PAD,
        y_baseline=baseline,
        max_width=key_w - 2 * CELL_PAD,
        latin_font=LATIN_FONT_BOLD,  # bold latin
        ar_font=AR_FONT_BOLD_NAME,   # bold arabic (registered above; falls back gracefully)
        kv_text=str(key_text),
        kv_size=font_size,
        leading=font_size + 2
    )

    # value (right cell) — normal color / weight
    cnv.setFillColor(black)
    draw_mixed_line(
        cnv=cnv,
        x_left=x + key_w + CELL_PAD,
        y_baseline=baseline,
        max_width=val_w - 2 * CELL_PAD,
        latin_font=font_latin,  # normal latin
        ar_font=font_ar,        # normal arabic
        kv_text=str(val_text),
        kv_size=font_size,
        leading=font_size + 2
    )


# ---------------- Main ----------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Path to input .xlsx")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    out_path = os.path.join("output", f"{datetime.now().strftime(r'%H%M%S_%d%m%Y')}.pdf")

    # Read Excel
    df = read_excel_preserve_display(excel_path)

    df = df.fillna("").astype(str)
    df.columns = [str(c).strip() for c in df.columns]

    # case-insensitive mapping to actual Excel headers
    lc_to_orig = {c.lower(): c for c in df.columns}
    keep_cols = [lc_to_orig[k.lower()] for k in FIELDS_TO_KEEP if k.lower() in lc_to_orig]

    # filter & reorder, fill empty cells, and force all data to string
    df = df[keep_cols].fillna("").astype(str)

    # make a full copy for title fields, also guaranteed to be str
    df_full = df.copy()

    # Drop title fields from the key-value printing DataFrame
    df = df[[c for c in df.columns if c.lower() not in [t.lower() for t in EXCLUSION_FIELDS]]]

    # Fonts: hard-coded auto-detection (no CLI arg)
    ar_font = register_arabic_font()  # Arabic-capable or fallback to Helvetica

    # Create output
    cnv = rl_canvas.Canvas(str(out_path), pagesize=(PAGE_WIDTH, PAGE_HEIGHT))
    cnv.setFillColor(black)

    # Calculate two column widths and positions
    content_width = (PAGE_WIDTH - margin_l - margin_r) * 2.90/5.0
    col_width = (content_width - col_gap) / 2.0              # two columns inside that block
    # anchor the 2-column block to the RIGHT side of the page
    block_left_x = PAGE_WIDTH - margin_r - content_width
    left_x = block_left_x
    right_x = block_left_x + col_width + col_gap

    for idx, row in df.iterrows():
        # Build title from the original full DF so we still have region/city/hay/pca
        row_full = df_full.iloc[idx]
        title_text = f"{row_full['Region']} – {row_full['City']} – {row_full['Hay']} – {row_full['PCA']}"

        # Use mixed-script renderer for title
        title_top_padding = 40  # custom padding from top edge
        y_title_top = PAGE_HEIGHT - margin_t - title_top_padding
        cnv.setFillColorRGB(98/255, 28/255, 154/255)  # #621c9a
        
        y_after_title = draw_mixed_line(
            cnv=cnv,
            x_left=margin_l,
            y_baseline=y_title_top,
            max_width=PAGE_WIDTH - margin_l - margin_r,
            latin_font=LATIN_FONT,
            ar_font=ar_font,
            kv_text=title_text,
            kv_size=TITLE_FONT_SIZE,
            leading=TITLE_FONT_SIZE + 6
        )
        cnv.setFillColor(black)  # reset for body

        # Add spacing before key-value columns start
        title_to_columns_gap = 2.5
        y_left = y_after_title - title_to_columns_gap
        y_right = y_after_title - title_to_columns_gap
        col_toggle = "left"

        # equal widths for key and value cells
        key_w = col_width / 2.0
        val_w = col_width / 2.0

        # how many rows fit in one column below the title
        available_height = min(y_left, y_right) - margin_b
        rows_per_col = max(1, int((available_height + row_gap) // (cell_height + row_gap)))

        # split list EQUALLY between LEFT and RIGHT (odd extra goes LEFT)
        all_keys = list(df.columns)
        n = len(all_keys)

        max_rows = rows_per_col
        left_n = min(max_rows, (n + 1) // 2)
        right_n = min(max_rows, n - left_n)
        col1_items = all_keys[:left_n]
        col2_items = all_keys[left_n:left_n + right_n]
        # (any leftover beyond 2 * max_rows is omitted to keep everything on a single page)

        # draw LEFT column
        y = y_left
        for kv_idx, key in enumerate(col1_items):
            raw_val = row[key]
            val = _normalize_value(raw_val)
            draw_kv_row(
                cnv, left_x, y, cell_height,
                key_text=key, val_text=val,
                key_w=key_w, val_w=val_w,
                font_latin=LATIN_FONT, 
                font_ar=ar_font, 
                font_size=KV_FONT_SIZE,
                row_index=kv_idx
            )
            y -= (cell_height + row_gap)
        y_left = y

        # draw RIGHT column
        y = y_right
        for kv_idx, key in enumerate(col2_items):
            raw_val = row[key]
            val = _normalize_value(raw_val)
            draw_kv_row(
                cnv, right_x, y, cell_height,
                key_text=key, val_text=val,
                key_w=key_w, val_w=val_w,
                font_latin=LATIN_FONT, 
                font_ar=ar_font, 
                font_size=KV_FONT_SIZE,
                row_index=kv_idx
            )
            y -= (cell_height + row_gap)
        y_right = y

        # Horizontal bounds for the image area
        img_left_x  = margin_l
        img_right_x = block_left_x - img_gap - img_pad
        img_width   = max(0, img_right_x - img_left_x)

        # Vertical bounds (same top as columns; bottom at margin)
        img_top    = y_after_title - title_to_columns_gap
        img_bottom = margin_b
        img_height = max(0, img_top - img_bottom)

        # Try to load the image from IMAGES FOLDER
        row_full = df_full.iloc[idx]  # original row with all fields
        img_name = str(row_full.get('image', '')).strip() if hasattr(row_full, "get") else ""
        img_path = os.path.join(IMAGES_PATH, img_name) 

        if img_width > 0 and img_height > 0 and img_name and os.path.exists(img_path):
            reader = ImageReader(str(img_path))
            iw, ih = reader.getSize()

            # scale to fit (preserve aspect ratio), anchored to start at the same top as the table
            avail_w = img_width
            avail_h = img_top - img_bottom  # don't cross the title area
            scale = min(avail_w / iw, avail_h / ih)
            draw_w = iw * scale
            draw_h = ih * scale

            # horizontally center within the left block, TOP‑ALIGN to the table start
            draw_x = img_left_x + (avail_w - draw_w) / 2
            draw_y = img_top - draw_h  # top-aligned so it never overlaps the title
            # clamp just in case
            if draw_y < img_bottom:
                draw_y = img_bottom

            cnv.drawImage(reader, draw_x, draw_y, width=draw_w, height=draw_h, mask='auto')
        # --- end LEFT-SIDE IMAGE ---
        
        cnv.showPage()
        
    cnv.save()
    print(f"PDF output saved to --> {out_path}.")


if __name__ == "__main__":
    main()
